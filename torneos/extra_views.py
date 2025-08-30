from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .models import Partido, EventoPartido, Jugador, Equipo
from openpyxl import load_workbook

class ImportJugadoresView(APIView):
    permission_classes = [permissions.IsAdminUser]

    def post(self, request, format=None):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error':'No file provided'}, status=400)
        try:
            wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': str(e)}, status=400)

        headers = [ (cell.value or '').strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1)) ]
        try:
            col_nombre = headers.index('nombre')
            col_equipo = headers.index('equipo')
        except ValueError:
            return Response({'error': 'Encabezados requeridos: Nombre, Equipo'}, status=400)

        created, errors = 0, []
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            nombre = (row[col_nombre].value or '').strip()
            equipo_nom = (row[col_equipo].value or '').strip()
            if not nombre or not equipo_nom:
                errors.append({'row': idx, 'error':'Falta nombre o equipo'})
                continue
            equipo = Equipo.objects.filter(nombre__iexact=equipo_nom).first()
            if not equipo:
                errors.append({'row': idx, 'error': f'Equipo "{equipo_nom}" no encontrado'})
                continue
            Jugador.objects.create(nombre=nombre, equipo=equipo)
            created += 1
        return Response({'created': created, 'errors': errors})

class TablaGoleadoresView(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self, request):
        data = []
        for j in Jugador.objects.select_related('equipo').all():
            goles_eventos = EventoPartido.objects.filter(jugador=j, tipo='gol').count()
            total = j.goles + goles_eventos
            if total>0:
                data.append({'jugador': j.nombre, 'equipo': j.equipo.nombre, 'goles': total})
        data.sort(key=lambda x: -x['goles'])
        return Response(data)

class TablaFairPlayView(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self, request):
        equipos = Equipo.objects.all()
        data = []
        for e in equipos:
            amarillas = EventoPartido.objects.filter(jugador__equipo=e, tipo='amarilla').count()
            rojas = EventoPartido.objects.filter(jugador__equipo=e, tipo='roja').count()
            puntos = amarillas*1 + rojas*3
            data.append({'equipo': e.nombre, 'amarillas': amarillas, 'rojas': rojas, 'puntos': puntos})
        data.sort(key=lambda x: x['puntos'])
        return Response(data)

class TablaSuspensionesView(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self, request):
        data = []
        for j in Jugador.objects.select_related('equipo').all():
            amarillas = EventoPartido.objects.filter(jugador=j, tipo='amarilla').count() + j.amarillas
            rojas = EventoPartido.objects.filter(jugador=j, tipo='roja').count() + j.rojas
            fechas_susp = (amarillas // 4) + rojas * 1
            if fechas_susp > 0:
                data.append({'jugador': j.nombre, 'equipo': j.equipo.nombre, 'amarillas': amarillas, 'rojas': rojas, 'fechas_susp': fechas_susp})
        return Response(data)

class TablaPosicionesView(APIView):
    permission_classes = [permissions.AllowAny]
    def get(self, request):
        result = {}
        categorias = Partido.objects.values_list('categoria__id','categoria__nombre').distinct()
        for cid, cname in categorias:
            equipos = Equipo.objects.filter(categoria_id=cid)
            tabla = { e.id: {'equipo': e.nombre, 'pj':0,'pg':0,'pe':0,'pp':0,'gf':0,'gc':0,'pts':0} for e in equipos }
            for p in Partido.objects.filter(categoria_id=cid, terminado=True).select_related('equipo_local','equipo_visitante'):
                gl = EventoPartido.objects.filter(partido=p, jugador__equipo=p.equipo_local, tipo='gol').count()
                gv = EventoPartido.objects.filter(partido=p, jugador__equipo=p.equipo_visitante, tipo='gol').count()
                tabla[p.equipo_local.id]['pj'] += 1
                tabla[p.equipo_visitante.id]['pj'] += 1
                tabla[p.equipo_local.id]['gf'] += gl; tabla[p.equipo_local.id]['gc'] += gv
                tabla[p.equipo_visitante.id]['gf'] += gv; tabla[p.equipo_visitante.id]['gc'] += gl
                if gl > gv:
                    tabla[p.equipo_local.id]['pg'] += 1; tabla[p.equipo_local.id]['pts'] += 3
                    tabla[p.equipo_visitante.id]['pp'] += 1
                elif gv > gl:
                    tabla[p.equipo_visitante.id]['pg'] += 1; tabla[p.equipo_visitante.id]['pts'] += 3
                    tabla[p.equipo_local.id]['pp'] += 1
                else:
                    tabla[p.equipo_local.id]['pe'] += 1; tabla[p.equipo_visitante.id]['pe'] += 1
                    tabla[p.equipo_local.id]['pts'] += 1; tabla[p.equipo_visitante.id]['pts'] += 1
            sorted_tabla = sorted(tabla.values(), key=lambda x: (-x['pts'], -(x['gf']-x['gc'])))
            result[cname] = sorted_tabla
        return Response(result)
